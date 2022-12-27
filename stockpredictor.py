from openpyxl import load_workbook
import sys
import numpy as np
from scipy.optimize import minimize
import statistics


"""
Input:
     - filename: str of an excel file
Output data:
    - market_prices: [100, 200, ..,] len(marker_prices) = num_months
    - monthly_data: [{analyst_code: target_price, ...},{...}, ...] len(monthly_data) = num_months,
    - analysts: set {analyst_code, analyst_code, ...}
"""
def GetData(filename):
    workbook = load_workbook(filename, data_only=True)
    ws = workbook["Sheet1"]

    source_data = ws['B6':'BK200']
    mkt_price_data = ws['B3':'BK3'][0]

    analyst_col = 0
    monthly_data = []
    mkt_prices = []
    analysts = set()

    while analyst_col < len(source_data[0])-1:
        current_month_data = dict()
        for row in range(len(source_data)):
            cell_value = source_data[row][analyst_col].value
            if cell_value != 0 and not cell_value is None:
                current_month_data[cell_value] = source_data[row][analyst_col+1].value
                analysts.add(cell_value)
        monthly_data.append(current_month_data)
        mkt_prices.append(mkt_price_data[analyst_col+1].value)
        analyst_col+=2

    return monthly_data, mkt_prices, analysts

"""
Input:
    - monthly_data: [{analyst_code: target_price, ...},{...}, ...] len(monthly_data) = num_months,
    - analysts: set {analyst_code, analyst_code, ...}
Output data:
    - analyst_data: [[100,200,300],[150,150,400],[...],] len(analyst_data) = num_analysts, len(analyst_data[0]) = num_months
"""
def Clean(monthly_data, analysts):
    def CleanSingleAnalyst(single_analyst_monthly_prections):
        for idx in range(len(single_analyst_monthly_prections)):
            price_target = single_analyst_monthly_prections[idx]
            if not price_target is None:
                continue

            def GetNextVal(idx):
                next_idx=idx+1
                while next_idx < len(single_analyst_monthly_prections):
                    val = single_analyst_monthly_prections[next_idx]
                    if not val is None:
                        return val, next_idx-idx+1
                    next_idx+=1
                return None, None

            next_val, next_distance = GetNextVal(idx)

            #no previous, yes next
            if idx==0 and not next_val is None:
                single_analyst_monthly_prections[idx] = next_val
            #yes previous, yes next
            elif idx != 0 and not next_val is None:
                prev_val = single_analyst_monthly_prections[idx-1]
                single_analyst_monthly_prections[idx] = prev_val + ((next_val-prev_val)/next_distance)
            #yes previous, no next
            elif idx != 0 and next_val is None:
                prev_val=single_analyst_monthly_prections[idx-1]
                single_analyst_monthly_prections[idx] = prev_val
        return single_analyst_monthly_prections

    sorted_analysts = sorted(list(analysts))
    cleaned_data = []
    for analyst in sorted_analysts:
        single_analyst_monthly_prections = []
        contiuous_none = 0
        previous_none = False
        for curr_month in monthly_data:
            if not analyst in curr_month or curr_month[analyst] is None or curr_month[analyst] == '@NA':
                single_analyst_monthly_prections.append(None)
                if previous_none:
                    contiuous_none+=1
                    if contiuous_none > 5:
                        break
                else:
                    contiuous_none=1
                previous_none=True
            else:
                single_analyst_monthly_prections.append(curr_month[analyst])
                previous_none = False
        if contiuous_none <= 5:
            single_analyst_monthly_prections = CleanSingleAnalyst(single_analyst_monthly_prections) if contiuous_none > 0 else single_analyst_monthly_prections
            cleaned_data.append(single_analyst_monthly_prections)
    return cleaned_data

"""
Input:
    - analyst_data: [[100,200,300],[150,150,400],[...],] len(analyst_data) = num_analysts, len(analyst_data[0]) = num_months
    - market_prices: [100, 200, ..,] len(marker_prices) = num_months
Output:
    - learned_weights: [float, ..] len = num_analysts
    - average_month, standard_deviation: floats
"""
def LearnWeights(analyst_data, mkt_prices):
    num_analysts = len(analyst_data)
    analyst_weights = [1/num_analysts for i in range (num_analysts)]
    analyst_weights = np.array(analyst_weights)
    
    def scoring_function(analyst_weights):
        return 1

    result = minimize(scoring_function)



"""
Input:
    - leanrned_weights: [float, ..] len = num_analysts
    - target_prices: [float, ...] len = num_analysts
Output:
    - predicted price floats
"""
def Predict():
    pass

filename = sys.argv[1]
analyst_data, mkt_prices, analysts = GetData(filename)

cleaned_data = Clean(analyst_data, analysts)

training_data = []
for row in cleaned_data:
    training_data.append(row[:-6])


leanrned_weights, average_months, stdv = LearnWeights(training_data, market_prices)
#
# predicted_price = Predict(leanrned_weights, analyst_data[-1])
#
# print("predicted price", predicted_price)
# print("estimated mo